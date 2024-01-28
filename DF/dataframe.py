import os
dir = os.path.dirname(os.path.abspath(__file__))
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

class DataFrame:
    def __init__(self, name, datas):
        self.__libro = openpyxl.Workbook()
        self.__hoja = self.__libro.active
        self.__border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.__bold = Font(bold=True)
        self.__row = 3
        self.__name = name
        self.__datas = datas
        self.__length = len(list(datas[0].keys()))
        self.__success = self.generateFrame()

    def getDirectory(self):
        return f"{dir}/{self.__name}.xlsx"

    def getSuccess(self):
        return self.__success

    def setTitles(self, titles):
        for i in range(len(titles)):
            cell = self.__hoja.cell(2, i+2, titles[i])
            cell.font = self.__bold
            cell.border = self.__border

    def setRow(self, register):
        for i in range(len(register)):
            cell = self.__hoja.cell(self.__row, i+2, register[i])
            cell.border = self.__border
        self.__row += 1

    def autoSize(self):
        for i in range(self.__length):
            max_length = 0
            for cell in list(self.__hoja.columns)[i+1]:
                if cell.value:
                    max_length = (
                        len(str(cell.value))
                        if len(str(cell.value)) > max_length
                        else max_length
                    )
            adjusted_width = max_length+2
            self.__hoja.\
            column_dimensions[get_column_letter(i+2)].\
            width = adjusted_width

    def saveFrame(self):
        self.__libro.save(f'{dir}/{self.__name}.xlsx')

    def deleteFrame(self):
        try:
            os.remove(f"{dir}/{self.__name}.xlsx")
            return True
        except Exception as ex:
            print( "-> Ocurrio un error al intentar "\
                  f"eliminar el dataframe: {str(ex)}.")
            return False

    def generateFrame(self):
        try:
            self.setTitles(list(self.__datas[0].keys()))
            for data in self.__datas:
                self.setRow(list(data.values()))
            self.autoSize()
            self.saveFrame()
            return True
        except Exception as ex:
            print( "-> Ocurrio un error al intentar "\
                  f"generar el dataframe: {str(ex)}.")
            return False